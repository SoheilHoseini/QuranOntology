# Code for building an ontology with class names in an excel file.
# Contributed by Soheil Hoseini - CE student at IUST - Iran, Tehran


from types import NoneType
import openpyxl # Reading an excel file using Python
import os # Used for converting the txt file to owl file

# Converts the Persian class names to .owl format
def ModifyEntityNameForOnto(className):
    nameList = list(className.split())
    
    #Remove paranthesis
    for i in range(len(nameList)):
        if nameList[i] == "(" or nameList[i] == ")":
            print(className)
            nameList[i] = "" 
            
        tmp = nameList[i]
        ans = ""
        
        for j in range(len(tmp)):
            
            if tmp[j] == "(" and tmp[j + 1] == "ع":
                ans += "_"
                continue
            
            if tmp[j] == "(" and tmp[j + 1] == "ص":
                ans += "_"
                continue
            
            if tmp[j] == "(" and tmp[j + 1] == "س":
                ans += "_"
                continue
            
            if tmp[j] != "(" and tmp[j] != ")":
                ans += tmp[j] 
                 
        nameList[i] = ans    
            
    return "_".join(nameList)


# Takes a class name and adds it the ontology
def AddClassToOntology(className, fileName):
    
    # Class declaration
    fileName.write("\n    <Declaration>\n")
    middleTxt = '        <Class IRI="#' + ModifyEntityNameForOnto(className) + '"/>'
    fileName.write(middleTxt)
    fileName.write("\n    </Declaration>")


# Takes a relation name and adds it the ontology
def AddClassObjPropToOntology(relation_name, domain_class, range_class, file_name, object_property_type):
    
    # Declaration of the relation
    file_name.write("\n    <Declaration>\n")
    middleTxt = '        <ObjectProperty IRI="#' + ModifyEntityNameForOnto(relation_name) + '"/>'
    file_name.write(middleTxt)
    file_name.write("\n    </Declaration>")
    
    # Determine the type of the object property
    if object_property_type != None:
        objPropStr = ObjectPropertyTypeString(object_property_type)
        
        startText = "\n    <" + objPropStr + ">\n"
        file_name.write(startText)
        
        middleTxt = '        <ObjectProperty IRI="#' + ModifyEntityNameForOnto(relation_name) + '"/>'
        file_name.write(middleTxt)
        
        endingText = "\n    </" + objPropStr + ">"
        file_name.write(endingText)
    
    # Determine domains of the object properties
    file_name.write("\n    <ObjectPropertyDomain>\n")
    
    middleTxt = '        <ObjectProperty IRI="#' + ModifyEntityNameForOnto(relation_name) + '"/>'
    file_name.write(middleTxt)
    
    middleTxt2 = '\n        <Class IRI="#' + ModifyEntityNameForOnto(domain_class) + '"/>'
    file_name.write(middleTxt2)
    
    file_name.write("\n    </ObjectPropertyDomain>")
    
    # Determine ranges of the object properties
    file_name.write("\n    <ObjectPropertyRange>\n")
    
    middleTxt = '        <ObjectProperty IRI="#' + ModifyEntityNameForOnto(relation_name) + '"/>'
    file_name.write(middleTxt)
    
    middleTxt2 = '\n        <Class IRI="#' + ModifyEntityNameForOnto(range_class) + '"/>'
    file_name.write(middleTxt2)
    
    file_name.write("\n    </ObjectPropertyRange>")
    
    
# Takes a object property type and returns the corresponding string of it
def ObjectPropertyTypeString(objectProType):
    
    objectProType = objectProType.lower()#convert to lower case for case insensitive comparison
    
    switcher = {
        "functional" : "FunctionalObjectProperty",
        "inverse functional" : "InverseFunctionalObjectProperty",
        "symmetric" : "SymmetricObjectProperty",
        "asymmetric" : "AsymmetricObjectProperty",
        "transitive" : "TransitiveObjectProperty",
        "reflexive" : "ReflexiveObjectProperty",
        "irreflexive" : "IrreflexiveObjectProperty",
        "Functional" : "FunctionalObjectProperty",
        "Inverse Functional" : "InverseFunctionalObjectProperty",
        "Symmetric" : "SymmetricObjectProperty",
        "Asymmetric" : "AsymmetricObjectProperty",
        "Transitive" : "TransitiveObjectProperty",
        "Reflexive" : "ReflexiveObjectProperty",
        "Irreflexive" : "IrreflexiveObjectProperty"
    }
    
    return switcher.get(objectProType)


# Take a class name and its father and add it to the ontology
def AddSubclasses(fatherClass, subclass , fileName):
    fileName.write("\n    <SubClassOf>\n")
    middleTxt1 = '        <Class IRI="#' + ModifyEntityNameForOnto(subclass) + '"/>'
    fileName.write(middleTxt1)
    middleTxt2 = '\n        <Class IRI="#' + ModifyEntityNameForOnto(fatherClass) + '"/>'
    fileName.write(middleTxt2)
    fileName.write("\n    </SubClassOf>")
 
 
# Take a class name and its individuals and add it to the ontology
def AddIndividualsToOntology(className, individualName, fileName): 
    
    # Declaration of the individual
    fileName.write("\n    <Declaration>\n")
    middleTxt = '        <NamedIndividual IRI="#' + ModifyEntityNameForOnto(individualName) + '"/>'
    fileName.write(middleTxt)
    fileName.write("\n    </Declaration>")
    
    # Cretae connection between individual and its referense class
    fileName.write("\n    <ClassAssertion>\n")
    middleTxt2 = '        <Class IRI="#' + ModifyEntityNameForOnto(className) + '"/>'
    fileName.write(middleTxt2)
    middleTxt3 = '\n        <NamedIndividual IRI="#' + ModifyEntityNameForOnto(individualName) + '"/>'
    fileName.write(middleTxt3)
    fileName.write("\n    </ClassAssertion>")


# Adds the initializing tags, needed for building an ontology
def InitializeOntology(ontoFile, ontologyName, annotation):
    
    ontoFile.write('<?xml version="1.0"?>\n')
    ontoFile.write('<Ontology xmlns="http://www.w3.org/2002/07/owl#"\n')
    ontoFile.write('     xml:base="http://www.sem0anticweb.org/rayan-tech/ontologies/2022/2/' + ontologyName + '"\n')
    ontoFile.write('     xmlns:rdf="http://www.w3.org/1999/02/22-rdf-syntax-ns#"\n')
    ontoFile.write('     xmlns:xml="http://www.w3.org/XML/1998/namespace"\n')
    ontoFile.write('     xmlns:xsd="http://www.w3.org/2001/XMLSchema#"\n')
    ontoFile.write('     xmlns:rdfs="http://www.w3.org/2000/01/rdf-schema#"\n')
    ontoFile.write('     ontologyIRI="http://www.sem0anticweb.org/rayan-tech/ontologies/2022/2/' + ontologyName + '">\n')
    ontoFile.write('     <Prefix name="" IRI="http://www.sem0anticweb.org/rayan-tech/ontologies/2022/2/' + ontologyName + '"/>\n')
    ontoFile.write('     <Prefix name="owl" IRI="http://www.w3.org/2002/07/owl#"/>\n')
    ontoFile.write('     <Prefix name="rdf" IRI="http://www.w3.org/1999/02/22-rdf-syntax-ns#"/>\n')
    ontoFile.write('     <Prefix name="xml" IRI="http://www.w3.org/XML/1998/namespace"/>\n')
    ontoFile.write('     <Prefix name="xsd" IRI="http://www.w3.org/2001/XMLSchema#"/>\n')
    ontoFile.write('     <Prefix name="rdfs" IRI="http://www.w3.org/2000/01/rdf-schema#"/>\n')
    
    ontoFile.write('     <Annotation>\n')
    ontoFile.write('         <AnnotationProperty abbreviatedIRI="rdfs:comment"/>\n')
    ontoFile.write('         <Literal>' + annotation + '</Literal>\n')
    ontoFile.write('     </Annotation>')
    
    
# Adds final tags for creating the ontology
def FinalizeOntology(file):
    file.write("\n</Ontology>")
    file.write('\n\n\n<!--' + 'Contributed by Soheil Hoseini' + '-->')
    

# Take a list of classes and subclasses and creates a hierarchy ontology
def CreateHierarchicalOntologyProcess(hirarchy_onto_excelName, ontology_file, father_column, subclass_column, excel_rel_path):
    
    allClassesList = []
    
    # Load classes and subclasses excel
    ontologySubclasses = openpyxl.load_workbook(excel_rel_path + hirarchy_onto_excelName + ".xlsx")
    subclassList = ontologySubclasses.active 


    # Iterate file of subclasses and declare each subclasses as a class in the ontology
    for i in range(2, subclassList.max_row + 1):
        
        fatherClassName  = subclassList.cell(row = i, column = father_column).value
        subclassName = subclassList.cell(row = i, column = subclass_column).value

        if (fatherClassName not in allClassesList):
            AddClassToOntology(fatherClassName , ontology_file)
            allClassesList.append(fatherClassName)
        
        if (type(subclassName) == NoneType):
            continue 
        
        if (subclassName not in allClassesList):
            AddClassToOntology(subclassName, ontology_file)
            allClassesList.append(subclassName)

        

    # Create connection between father class and its subclasses
    for i in range(2, subclassList.max_row + 1):
        
        fatherClassName  = subclassList.cell(row = i, column = father_column).value
        subclassName = subclassList.cell(row = i, column = subclass_column).value
        
        if (type(subclassName) == NoneType):
            continue
            
        AddSubclasses(fatherClassName, subclassName, ontology_file)

# Takes a relation name and adds it the ontology
def AddIndvObjPropToOntology(relationName, domainIndv, rangeIndv, fileName, objectPropertyType):
    
    # Declaration of the object property
    fileName.write("    <Declaration>\n")
    middleTxt = '        <ObjectProperty IRI="#' + ModifyEntityNameForOnto(relationName) + '"/>'
    fileName.write(middleTxt)
    fileName.write("\n    </Declaration>")
    
    # Determine the type of the object property
    if objectPropertyType != None:
        objPropStr = ObjectPropertyTypeString(objectPropertyType)

        startText = "\n    <" + objPropStr + ">\n"
        fileName.write(startText)
        
        middleTxt = '        <ObjectProperty IRI="#' + ModifyEntityNameForOnto(relationName) + '"/>'
        fileName.write(middleTxt)
        
        endingText = "\n    </" + objPropStr + ">"
        fileName.write(endingText)
    
    # Determine domain and range of the object properties
    fileName.write("\n    <ObjectPropertyAssertion>\n")
    
    middleTxt = '        <ObjectProperty IRI="#' + ModifyEntityNameForOnto(relationName) + '"/>'
    fileName.write(middleTxt)
    
    middleTxt2 = '\n        <NamedIndividual IRI="#' + ModifyEntityNameForOnto(domainIndv) + '"/>'
    fileName.write(middleTxt2)
    
    middleTxt3 = '\n        <NamedIndividual IRI="#' + ModifyEntityNameForOnto(rangeIndv) + '"/>'
    fileName.write(middleTxt3)
    
    fileName.write("\n    </ObjectPropertyAssertion>\n") 
    

# Take ontology name and annotation from user
ontologyName = input("Enter the name of your ontology: \n") # There shouldn't be any spaces in the ontology name
annotation = input("Enter a brief annotation for your ontology: \n")

# Create a txt file for the ontology to write initializing owl tags to it
newOntology = open("txt files\\" + ontologyName + ".txt", 'w', encoding="utf8")

# Initialize the Ontology
InitializeOntology(newOntology, ontologyName, annotation)

all_classes_list = []
all_indvs_list = []


# TO DELETE BEGINNING
path = "xlsx files\\IndividualOntology\\Phase3\\" 
excel_file_name = "FatherChildList"

excel_file = openpyxl.load_workbook(path + excel_file_name + ".xlsx")
records_list = excel_file.active

all_entities = open("resources\\all_entities.txt", 'w', encoding="utf8")

for i in range(2, records_list.max_row + 1):
    
    root_class = records_list.cell(row = i, column = 3).value
    indv1 = records_list.cell(row = i, column = 2).value
    indv2 = records_list.cell(row = i, column = 1).value
    
    if root_class not in all_classes_list:
        all_classes_list.append(root_class)
        all_entities.write(root_class+"\n")
        AddClassToOntology(root_class, newOntology)
    
    if indv1 not in all_indvs_list:
        all_indvs_list.append(indv1)
        all_entities.write(indv1+"\n")
        AddIndividualsToOntology(root_class, indv1, newOntology)
        
    if indv2 not in all_indvs_list and root_class != indv2:
        all_indvs_list.append(indv2)
        all_entities.write(indv2+"\n")
        AddIndividualsToOntology(root_class, indv2, newOntology)    
        
        AddIndvObjPropToOntology("هست يك", indv1, indv2, newOntology, "transitive")

path2 = "xlsx files\\IndividualOntology\\Phase3\\"
rel_file_name = "RefinedRelations2" 

rels_excel_file = openpyxl.load_workbook(path2 + rel_file_name + ".xlsx")
rels_list = rels_excel_file.active

for i in range(2 , rels_list.max_row + 1):
    domain_indv = rels_list.cell(i, 1).value
    range_indv = rels_list.cell(i, 2).value
    obj_prop_type = rels_list.cell(i, 3).value
    obj_prop_name = rels_list.cell(i, 4).value
    
    if domain_indv not in all_indvs_list:
        all_indvs_list.append(domain_indv)
        
    if range_indv not in all_indvs_list:
        all_indvs_list.append(range_indv)
    
    AddIndvObjPropToOntology(obj_prop_name, domain_indv, range_indv, newOntology, obj_prop_type)    
# TO DELETE END


'''
# Load object properties excel with its path
ontologyObjProp = openpyxl.load_workbook("QuranObjectProperties.xlsx")
objPropList = ontologyObjProp.active

# Iterate through excel file of relations and add classes to ontology
for i in range(2, objPropList.max_row+1):
    domainClass = objPropList.cell(row=i, column=1).value
    rangeClass = objPropList.cell(row=i, column=2).value
    relationName = objPropList.cell(row=i, column=3).value
    objectPropertyType = objPropList.cell(row=i, column=4).value
    AddObjectPropToOntology(relationName, domainClass, rangeClass, newOntology, objectPropertyType)
'''  


'''
# Load Individuals
concept_indv_excel_name = "QuranConceptIndv"
ontologyIndv = openpyxl.load_workbook("xlsx files\\IndividualOntology\\" + concept_indv_excel_name + ".xlsx")
indvList = ontologyIndv.active

# Iterate file of individuals and add them to the ontology
for i in range(2, indvList.max_row+1):
    classNameForIndv = indvList.cell(row=i, column=2).value
    indvName = indvList.cell(row=i, column=1).value
    
    if(type(indvName) == NoneType):
        continue
    
    AddIndividualsToOntology(classNameForIndv, indvName, newOntology)
'''
all_entities.close()


# Add final tags to the ontology file
FinalizeOntology(newOntology)
newOntology.close()

# Convert the txt file to owl file
convertedFile = "txt files\\" + ontologyName + ".txt"
base = os.path.splitext(convertedFile)[0]
base = "owl" + base[3:] # Save the ontology in the "owl files" folder
os.rename(convertedFile, base + '.owl')

print('Congratulations! You have successfully built your new ontology.')

