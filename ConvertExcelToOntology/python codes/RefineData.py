# Refine input data to ontology to homogenize Persian and Arabic phonetics
# Contributed by Soheil Hoseini


from types import NoneType
import openpyxl# Reading an excel file using Python
import xlsxwriter # Create and fill xlsx file 

class RefineOntologyData:
    
    def __init__(self) -> None:
        
        self.modified_y_cnt = 0
        
        self.modified_half_space_cnt = 0
        self.modified_half_space_indices = list()
        
        self.modified_v_cnt = 0
        self.modified_v_indices = list()
        
        self.modified_h_cnt = 0
        self.modified_h_indices = list()
        
        self.modified_a_cnt  = 0
        self.modified_a_indices = list()
    
    def modify_enitiy_name(self, name, entity_row_indx):
        # در فایل پدر و فرزندی، "ی" فارسی هست و کیبورد رو میخونه برعکس فایل روابط
        if name == NoneType:
            print("None", name)
            
        names_list = list(name.split())
        
            
        file = open("resources\\SpecialCharacters.txt", 'r', encoding="utf8")
        half_space = file.read()
        file.close()
        for j in range(len(names_list)):
            
            word = names_list[j]
            ans = ""
            for k in range(len(word)):
                
                # Convet Arabic format to Persian
                if word[k] == "ى":
                    ans += "ی"
                    self.modified_y_cnt += 1
                
                # Remove half spaces from words
                elif word[k] == half_space:
                    if word[k-1] != "_":
                        ans += " "
                    
                    if k + 1 <= len(word) - 1 and word[k+1] != "_":
                        ans += " "
                        
                    self.modified_half_space_cnt += 1
                    self.modified_half_space_indices.append(entity_row_indx)
                
                
                # Convet Arabic format to Persian (Only for appearing at the end of words that are in the middle of phrases)
                # the word should not be at the end of the phrase
                elif word[k] == "ة" and not (j != (len(names_list) - 1) and k == (len(word) - 1)):
                    ans += "ه"
                    self.modified_h_cnt += 1
                    self.modified_h_indices.append(entity_row_indx)
                
                
                elif word[k] == "أ" and ((k == len(word) - 1) or k == 0):
                    ans += "ا"
                    self.modified_a_cnt += 1
                    self.modified_a_indices.append(entity_row_indx)
                
                
                elif word[k] == "ؤ" and ((k == len(word) - 1) or k == 0):
                    ans += "و"
                    self.modified_v_cnt += 1
                    self.modified_v_indices.append(entity_row_indx)
                
                
                elif word[k] == "ۀ":
                    ans += "ه"
                    self.modified_h_cnt += 1
                    self.modified_h_indices.append(entity_row_indx)
                
                
                else:
                    ans += word[k]
                
            
            names_list[j] = ans
        return " ".join(names_list)  
    

refine_engine = RefineOntologyData()
        
path = "xlsx files\\IndividualOntology\\Phase2\\" 
excel_file_name = "ScienceIndvs"

excel_file = openpyxl.load_workbook(path + excel_file_name + ".xlsx")
records_list = excel_file.active

empty_rows = list()


dest_xlsx_file = "ScienceIndvsPh4"
destin_path = "xlsx files\\IndividualOntology\\Phase4\\"
workbook = xlsxwriter.Workbook(destin_path + dest_xlsx_file + ".xlsx")
worksheet = workbook.add_worksheet() 


row = 1
# Add the original file headers
worksheet.write(0, 0, records_list.cell(1,1).value)
worksheet.write(0, 1, records_list.cell(1,2).value)
worksheet.write(0, 2, records_list.cell(1,3).value)

for i in range(2, records_list.max_row + 1):
    
    cell1 = records_list.cell(i, 1).value
    cell2 = records_list.cell(i, 2).value
    #cell3 = records_list.cell(i, 3).value
    #cell4 = records_list.cell(i, 4).value
    
    if(type(cell1) == NoneType or type(cell2) == NoneType):
        empty_rows.append(i)
        continue
    
    modified_cell1 = refine_engine.modify_enitiy_name(cell1, i)
    worksheet.write(row, 0, modified_cell1)
    
    modified_cell2 = refine_engine.modify_enitiy_name(cell2, i)
    worksheet.write(row, 1, modified_cell2)
    
    #modified_cell3 = refine_engine.modify_enitiy_name(cell3, i)
    #worksheet.write(row, 2, cell3)
    
    #modified_cell4 = refine_engine.modify_enitiy_name(cell4, i)
    #worksheet.write(row, 3, modified_cell4)
    
    row += 1

workbook.close()

print(empty_rows)

print("Congrats my friend! Your data has been successfully refined. ")
print(f"\nStats:\n      Modified ی counts: {refine_engine.modified_y_cnt}\n")
print(f"      Modified half spaces counts: {refine_engine.modified_half_space_cnt} => {refine_engine.modified_half_space_indices}\n")
print(f"      Modified ه counts: {refine_engine.modified_h_cnt} => {refine_engine.modified_h_indices}\n")
print(f"      Modified و counts: {refine_engine.modified_v_cnt} => {refine_engine.modified_v_indices}\n")
print(f"      Modified ا counts: {refine_engine.modified_a_cnt} => {refine_engine.modified_a_indices}\n")