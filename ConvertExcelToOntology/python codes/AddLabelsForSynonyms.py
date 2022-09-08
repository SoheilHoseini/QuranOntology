# Code to add labels to class names as their synonyms.
# Contributed by Soheil Hoseini


from asyncio.windows_events import NULL
from io import TextIOWrapper
import os
import string
from types import NoneType
import openpyxl # Reading an excel file using Python

class Synonyms:
    
    def __init__(self, file: TextIOWrapper) -> None:
        self.onto_file = file
        
        
    # Con
    # verts the Persian class names to .owl format
    def ModifyEntityNameForOnto(self, className):
        nameList = list(className.split())
        
        #Remove paranthesis
        for i in range(len(nameList)):
            if nameList[i] == "(" or nameList[i] == ")":
                print(className)
                nameList[i] = "" 
                
            tmp = nameList[i]
            ans = ""
            
            for j in range(len(tmp)):
                
                if tmp[j] == "(" and tmp[j + 1] == "ع":
                    ans += "_"
                    continue
                
                if tmp[j] == "(" and tmp[j + 1] == "ص":
                    ans += "_"
                    continue
                
                if tmp[j] == "(" and tmp[j + 1] == "س":
                    ans += "_"
                    continue
                
                if tmp[j] != "(" and tmp[j] != ")":
                    ans += tmp[j] 
                    
            nameList[i] = ans    
                
        return "_".join(nameList)


    def AppendAnnotationLabels(self, entity_name: string, synonym: string, final_onto: TextIOWrapper):
        final_onto.write("\n    <AnnotationAssertion>")
        final_onto.write("\n        <AnnotationProperty abbreviatedIRI=\"rdfs:label\"/>")
        final_onto.write(f"\n        <IRI>#{self.ModifyEntityNameForOnto(entity_name)}</IRI>")
        final_onto.write(f"\n        <Literal>{self.ModifyEntityNameForOnto(synonym)}</Literal>")
        final_onto.write("\n    </AnnotationAssertion>")


    def AppendAnnotLablesProcess(self, 
                                 syn_excel_path: string, 
                                 syns_excel_name: string, 
                                 final_onto_name: string, 
                                 final_onto_path: string,
                                 all_entities: list):
        
        
        finalOntology = open(final_onto_path + final_onto_name + ".txt", "w", encoding="utf8")
        not_availavle_entities = list()
        current_entities = list()
        
        ontology_synonyms = openpyxl.load_workbook(syn_excel_path + syns_excel_name + ".xlsx")
        synonyms_list = ontology_synonyms.active
        
        for line in self.onto_file:
                              
            if line == "</Ontology>\n" or line == "</Ontology>":

                # Iterate through excel file of the synonyms
                for i in range(2, synonyms_list.max_row + 1):
                    
                    entity_name = synonyms_list.cell(i, 1).value
                    synonym = synonyms_list.cell(i, 2).value

                    current_entities.append(entity_name)
                    
                    if (entity_name not in all_entities) and (entity_name not in not_availavle_entities):
                        not_availavle_entities.append(entity_name)
                        continue
                    self.AppendAnnotationLabels(entity_name, synonym, finalOntology)
                    
            else:
                finalOntology.write(line)
                
        self.FinalizeOntologyTags(finalOntology)
        #self.ConvertTxtToOwl(final_onto_name)
        
        print(f"All stored entities: {len(all_entities)}")
        print(f"We checked {len(current_entities)} names.\n", current_entities,"\n\n\n\n")
        print(f"{len(not_availavle_entities)} number of entities were not found!\n", not_availavle_entities)

    # Adds final tags for creating the ontology
    def FinalizeOntologyTags(self, file: TextIOWrapper):
        file.write("\n</Ontology>")
        #file.write('\n\n\n<!--' + 'Contributed by Soheil Hoseini' + '-->')
    
    
    def ConvertTxtToOwl(self, ontology_name: string):
        
        # Convert the txt file to owl file
        convertedFile = "txt files\\" + ontology_name + ".txt"
        base = os.path.splitext(convertedFile)[0]
        base = "owl" + base[3:] # Save the ontology in the "owl files" folder
        os.rename(convertedFile, base + '.owl')

        print('Congratulations! You have successfully built your new ontology.')


path = "txt files\\"
orig_ontology_name = "IndvOntoLabel4"
orig_ontology_file = open(path + orig_ontology_name + ".txt", "r", encoding="utf8")

all_entities_file = open("resources\\all_entities.txt", 'r', encoding="utf8")
all_entities = list()
for line in all_entities_file:
    entity = line[:len(line)-1]
    if entity != "\n" and entity != "" and entity != None and entity != NoneType and entity != NULL:
        all_entities.append(entity)

syn_path = "xlsx files\\IndividualOntology\\Phase3\\Vol4\\"
onto_syn_excel_name = "QuranConceptsSynonymsVol4"

final_ontology_path = "txt files\\"
final_ontology_name = "IndvOntoFinal4"

# Start the adding synonyms process
synonyms_engine = Synonyms(orig_ontology_file)
synonyms_engine.AppendAnnotLablesProcess(syn_path, onto_syn_excel_name, final_ontology_name, final_ontology_path, all_entities)

orig_ontology_file.close()


# Convert the txt file to owl file
convertedFile = "txt files\\" + final_ontology_name + ".txt"
base = os.path.splitext(convertedFile)[0]
base = "owl" + base[3:] # Save the ontology in the "owl files" folder
os.rename(convertedFile, base + '.owl')

print('Congratulations! You have successfully built your new ontology.')

