# Code for adding object properties between individuals.
# Contributed by Soheil Hoseini


import openpyxl # Reading an excel file using Python
import os # Used for converting the txt file to owl file
from types import NoneType

# Converts the Persian class names to .owl format
def ModifyEntityNameForOnto(className):
    nameList = list(className.split())
    
    #Remove paranthesis
    for i in range(len(nameList)):
        if nameList[i] == "(" or nameList[i] == ")":
            print(className)
            nameList[i] = "" 
            
        tmp = nameList[i]
        ans = ""
        
        for j in range(len(tmp)):
            
            if tmp[j] == "(" and tmp[j + 1] == "ع":
                ans += "_"
                continue
            
            if tmp[j] == "(" and tmp[j + 1] == "ص":
                ans += "_"
                continue
            
            if tmp[j] == "(" and tmp[j + 1] == "س":
                ans += "_"
                continue
            
            if tmp[j] != "(" and tmp[j] != ")":
                ans += tmp[j] 
                 
        nameList[i] = ans    
            
    return "_".join(nameList)


# Takes a relation name and adds it the ontology
def AddIndvObjPropToOntology(relationName, domainIndv, rangeIndv, fileName, objectPropertyType):
    
    # Declaration of the object property
    fileName.write("    <Declaration>\n")
    middleTxt = '        <ObjectProperty IRI="#' + ModifyEntityNameForOnto(relationName) + '"/>'
    fileName.write(middleTxt)
    fileName.write("\n    </Declaration>")
    
    # Determine the type of the object property
    if objectPropertyType != None:
        objPropStr = ObjectPropertyTypeString(objectPropertyType)

        startText = "\n    <" + objPropStr + ">\n"
        fileName.write(startText)
        
        middleTxt = '        <ObjectProperty IRI="#' + ModifyEntityNameForOnto(relationName) + '"/>'
        fileName.write(middleTxt)
        
        endingText = "\n    </" + objPropStr + ">"
        fileName.write(endingText)
    
    # Determine domain and range of the object properties
    fileName.write("\n    <ObjectPropertyAssertion>\n")
    
    middleTxt = '        <ObjectProperty IRI="#' + ModifyEntityNameForOnto(relationName) + '"/>'
    fileName.write(middleTxt)
    
    middleTxt2 = '\n        <NamedIndividual IRI="#' + ModifyEntityNameForOnto(domainIndv) + '"/>'
    fileName.write(middleTxt2)
    
    middleTxt3 = '\n        <NamedIndividual IRI="#' + ModifyEntityNameForOnto(rangeIndv) + '"/>'
    fileName.write(middleTxt3)
    
    fileName.write("\n    </ObjectPropertyAssertion>\n") 
    
    
# Takes a object property type and returns the corresponding string of it
def ObjectPropertyTypeString(objectProType):
    
    objectProType = objectProType.lower()#convert to lower case for case insensitive comparison
    
    switcher = {
        "functional" : "FunctionalObjectProperty",
        "inverse functional" : "InverseFunctionalObjectProperty",
        "symmetric" : "SymmetricObjectProperty",
        "asymmetric" : "AsymmetricObjectProperty",
        "transitive" : "TransitiveObjectProperty",
        "reflexive" : "ReflexiveObjectProperty",
        "irreflexive" : "IrreflexiveObjectProperty",
        "Functional" : "FunctionalObjectProperty",
        "Inverse Functional" : "InverseFunctionalObjectProperty",
        "Symmetric" : "SymmetricObjectProperty",
        "Asymmetric" : "AsymmetricObjectProperty",
        "Transitive" : "TransitiveObjectProperty",
        "Reflexive" : "ReflexiveObjectProperty",
        "Irreflexive" : "IrreflexiveObjectProperty"
    }
    
    return switcher.get(objectProType)


# Take a class name and its individuals and add it to the ontology
def AddIndividualsToOntology(className, individualName, fileName): 
    
    # Declaration of the individual
    fileName.write("\n    <Declaration>\n")
    middleTxt = '        <NamedIndividual IRI="#' + ModifyEntityNameForOnto(individualName) + '"/>'
    fileName.write(middleTxt)
    fileName.write("\n    </Declaration>")
    
    # Correlate the individual to its base class
    fileName.write("\n    <ClassAssertion>\n")
    middleTxt2 = '        <Class IRI="#' + ModifyEntityNameForOnto(className) + '"/>'
    fileName.write(middleTxt2)
    middleTxt3 = '\n        <NamedIndividual IRI="#' + ModifyEntityNameForOnto(individualName) + '"/>'
    fileName.write(middleTxt3)
    fileName.write("\n    </ClassAssertion>")


# Adds final tags for creating the ontology
def FinalizeOntology(file):
    file.write("\n</Ontology>")
    file.write('\n\n\n<!--' + 'Contributed by Soheil Hoseini' + '-->')
    

# Append the object properties to the ontolog
def AppendInvdsObjPropsProcess(finalOntology, objPropsExcelName):
    
    # Read the excel file of the object properties
    indvObjProp = openpyxl.load_workbook("xlsx files\\" + objPropsExcelName + ".xlsx")
    indvObjPropList = indvObjProp.active
    
    #print(indvObjPropList.max_row + 1)
    
    for i in range(2, indvObjPropList.max_row + 1):
        
        domainIndv = indvObjPropList.cell(row = i, column = 1).value
        rangeIndv = indvObjPropList.cell(row = i, column = 2).value
        relationName = indvObjPropList.cell(row = i, column = 3).value
        objectPropertyType = indvObjPropList.cell(row = i, column = 4).value
        #print("relationName:" ,relationName," domainIndv:", domainIndv,
        #      " rangeIndv:",rangeIndv," objectPropertyType:", objectPropertyType)
        AddIndvObjPropToOntology(relationName, domainIndv, rangeIndv, finalOntology, objectPropertyType)
            
    
def AppendInvdsProcess():
    pass

# Create a txt file for the ontology to write initializing owl tags to it
ontologyName = "IndiRelTest"
origOntology = open("txt files\\" + ontologyName + ".txt","r", encoding="utf8")

finalOntoName = "FinalIndiRelTest"
finalOntology = open("txt files\\" + finalOntoName + ".txt", "w", encoding="utf8")


# 
objPropsExcelName = "IndvObjPropTest-Delete"
for line in origOntology:
    if line != "</Ontology>\n":
        finalOntology.write(line)
    else:
        AppendInvdsObjPropsProcess(finalOntology, objPropsExcelName)
        AppendInvdsProcess()
        FinalizeOntology(finalOntology)


origOntology.close()
finalOntology.close()

# Convert the txt file to owl file
convertedFile = "txt files\\" + finalOntoName + ".txt"
base = os.path.splitext(convertedFile)[0]
base = "owl" + base[3:] # Save the ontology in the "owl files" folder
os.rename(convertedFile, base + '.owl')

print('Congratulations! You have successfully built your new ontology.')

