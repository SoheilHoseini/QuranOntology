from ast import Pass
import openpyxl # Reading an excel file using Python
import xlsxwriter # Create and fill xlsx file 

hirarchy1_excel_name = "hierarchy1-Delete"
hirarchy2_excel_name = "hierarchy2-Delete"


# Read the excel file of the hirarchies
hirarchy1 = openpyxl.load_workbook("xlsx files\\" + hirarchy1_excel_name + ".xlsx")
hirarchy1_list = hirarchy1.active

hirarchy2 = openpyxl.load_workbook("xlsx files\\" + hirarchy2_excel_name + ".xlsx")
hirarchy2_list = hirarchy2.active

hierarchy1_records = list()
hierarchy2_records = list() 

all_entities = list()
for i in range(2, hirarchy1_list.max_row + 1):  
    father_class = hirarchy1_list.cell(row = i, column = 1).value
    child_class = hirarchy1_list.cell(row = i, column = 2).value
    hierarchy1_records.append((father_class, child_class))
    
    # Add father class and child class to a list seperately
    if father_class not in all_entities:
        all_entities.append(father_class)
        
    if child_class not in all_entities:
        all_entities.append(child_class)

    
for i in range(2, hirarchy2_list.max_row + 1):  
    father_class = hirarchy2_list.cell(row = i, column = 1).value
    child_class = hirarchy2_list.cell(row = i, column = 2).value        
    hierarchy2_records.append((father_class, child_class))
    
    # Add father class and child class to a list seperately
    if father_class not in all_entities:
        all_entities.append(father_class)
        
    if child_class not in all_entities:
        all_entities.append(child_class)
        
        
common_records = list()
common_records_list = list()

for i in range(len(hierarchy1_records)):
    for j in range(len(hierarchy2_records)):
        if hierarchy1_records[i] == hierarchy2_records[j]:
            common_records.append({hierarchy1_records[i]:[i+2,j+2]})
            common_records_list.append(hierarchy1_records[i])
        


# # Create a new xlsx file combined of the last two 
# workbook = xlsxwriter.Workbook("xlsx files\\" + "CombinedHierarchies.xlsx")
# worksheet = workbook.add_worksheet() 

# row = 1
# for i in range(len(hierarchy1_records)):
#     if hierarchy1_records[i] in common_records_list:
#         continue
#     worksheet.write(row, 0, hierarchy1_records[i][0])
#     worksheet.write(row, 1, hierarchy1_records[i][1])

#     row += 1

# for i in range(len(hierarchy2_records)):
    
#     worksheet.write(row, 0, hierarchy2_records[i][0])
#     worksheet.write(row, 1, hierarchy2_records[i][1])

#     row += 1
      
# workbook.close()


# # Add common records to a new excel file
# workbook2 = xlsxwriter.Workbook("xlsx files\\" + "CommonRecords.xlsx")
# worksheet2 = workbook2.add_worksheet()
# row = 1 
# for i in range(len(common_records)):
#     for key , value in common_records[i].items():
#         worksheet2.write(row, 0, key[0])
#         worksheet2.write(row, 1, key[1])
#         worksheet2.write(row, 2, value[0])
#         worksheet2.write(row, 3, value[1])
#     row += 1 
    
# workbook2.close()



# Read the excel file of the relations
rels_file_name = "relations-Delete"
relations = openpyxl.load_workbook("xlsx files\\" + rels_file_name + ".xlsx")
relations_list = relations.active

differences = list()

for i in range(2, relations_list.max_row + 1):
    domain = relations_list.cell(row = i, column = 1).value
    range_class = relations_list.cell(row = i, column = 2).value
    
    if domain not in all_entities:
        differences.append({domain:i})
    
    if range_class not in all_entities:
        differences.append({range_class:i})

print(len(differences))        
print(differences)


# Add different records to a new excel file
workbook3 = xlsxwriter.Workbook("xlsx files\\" + "DifferentRecords.xlsx")
worksheet3 = workbook3.add_worksheet()
row = 1 
for j in range(len(differences)):
    for key , value in differences[j].items():
        worksheet3.write(row, 0, key)
        worksheet3.write(row, 1, value)
    row += 1 
    
workbook3.close()